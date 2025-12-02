#!/usr/bin/env python3
"""
CSV/XLSX to PostgreSQL Import Script

Import data from CSV/XLSX files to zen_contacts table without re-scraping.
Only validates phone numbers via WAHA API.

Dedupe Strategy: name + street + city + country_code (100% unique in test data)

Usage:
    python toolkit/import_to_db.py result/Argentina.csv
    python toolkit/import_to_db.py result/Report.xlsx --sheet Argentina
    python toolkit/import_to_db.py file.csv --skip-waha --dry-run
"""

import argparse
import csv
import hashlib
import logging
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

import phonenumbers
from phonenumbers import NumberParseException

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None

# Project imports
from whatsapp_validator import WhatsAppValidator
from database_writer import DatabaseWriter, DatabaseConfig, create_database_writer


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class DataImporter:
    """Import CSV/XLSX data to PostgreSQL with deduplication and WAHA validation."""
    
    # Expected CSV columns
    EXPECTED_COLUMNS = ['name', 'street', 'city', 'country_code', 'phone_number']
    
    # Summary sheet indicators (skip these)
    SUMMARY_INDICATORS = ['total', 'sheet name', 'unique countries', 'summary']
    
    def __init__(self, db_writer: DatabaseWriter, waha_validator: Optional[WhatsAppValidator] = None,
                 dry_run: bool = False, batch_size: int = 100):
        self.db = db_writer
        self.waha = waha_validator
        self.dry_run = dry_run
        self.batch_size = batch_size
        
        # Stats
        self.stats = {
            'total_rows': 0,
            'inserted': 0,
            'merged': 0,
            'skipped': 0,
            'waha_validated': 0,
            'errors': 0
        }
    
    def generate_dedupe_key(self, row: Dict[str, Any]) -> str:
        """Generate deduplication key from name + street + city + country."""
        name = str(row.get('name', '')).lower().strip()
        street = str(row.get('street', '')).lower().strip()
        city = str(row.get('city', '')).lower().strip()
        country = str(row.get('country_code', '')).upper().strip()[:2]
        
        key_str = f"{name}|{street}|{city}|{country}"
        return key_str
    
    def generate_source_link(self, dedupe_key: str) -> str:
        """Generate unique source_link from dedupe key."""
        hash_val = hashlib.md5(dedupe_key.encode('utf-8')).hexdigest()[:16]
        return f"import:{hash_val}"
    
    def normalize_phone(self, phone: str, country_code: str = None) -> Optional[str]:
        """Normalize phone number to E.164 format."""
        if not phone:
            return None
        
        phone = str(phone).strip()
        
        # Remove common formatting
        phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
        
        # Handle float format from Excel (e.g., "6592282000.0")
        if '.' in phone and phone.endswith('.0'):
            phone = phone.replace('.0', '')
        
        try:
            # Try parsing with + prefix
            if phone.startswith('+'):
                parsed = phonenumbers.parse(phone, None)
            elif country_code:
                parsed = phonenumbers.parse(phone, country_code.upper())
            else:
                # Try adding + prefix for international numbers
                parsed = phonenumbers.parse('+' + phone, None)
            
            if phonenumbers.is_valid_number(parsed):
                return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
        except NumberParseException:
            pass
        
        return None
    
    def check_existing(self, name: str, street: str, city: str, country_code: str) -> Optional[Dict]:
        """Check if business already exists in database."""
        if self.dry_run:
            return None
        
        conn = self.db.pool.getconn()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT id, source_link, emails, phones, whatsapp,
                       social_facebook, social_instagram, business_website
                FROM zen_contacts 
                WHERE LOWER(business_name) = LOWER(%s) 
                  AND LOWER(COALESCE(street, '')) = LOWER(%s)
                  AND LOWER(city) = LOWER(%s) 
                  AND country_code = %s
                LIMIT 1
            """, (name, street, city, country_code.upper()[:2]))
            
            row = cursor.fetchone()
            if row:
                return {
                    'id': row[0],
                    'source_link': row[1],
                    'emails': row[2] or [],
                    'phones': row[3] or [],
                    'whatsapp': row[4] or [],
                    'social_facebook': row[5],
                    'social_instagram': row[6],
                    'business_website': row[7]
                }
            return None
        finally:
            self.db.pool.putconn(conn)
    
    def merge_arrays(self, existing: List[str], new_items: List[str]) -> List[str]:
        """Merge arrays without duplicates."""
        if not existing:
            existing = []
        if not new_items:
            return existing
        
        # Normalize and dedupe
        seen = set(str(x).lower().strip() for x in existing if x)
        result = list(existing)
        
        for item in new_items:
            item_str = str(item).strip()
            if item_str and item_str.lower() not in seen:
                result.append(item_str)
                seen.add(item_str.lower())
        
        return result
    
    def prepare_row(self, row: Dict[str, Any], validate_whatsapp: bool = True) -> Dict[str, Any]:
        """Prepare row data for database insertion/update."""
        country_code = str(row.get('country_code', 'XX')).upper().strip()[:2]
        
        # Normalize phone
        phone_raw = row.get('phone_number', '')
        phone_normalized = self.normalize_phone(phone_raw, country_code)
        
        # Prepare arrays
        emails = []
        if row.get('email'):
            emails = [str(row['email']).strip()]
        if row.get('emails'):
            if isinstance(row['emails'], str):
                emails.extend([e.strip() for e in row['emails'].split(';') if e.strip()])
            elif isinstance(row['emails'], list):
                emails.extend(row['emails'])
        
        phones = []
        if phone_normalized:
            phones = [phone_normalized]
        
        # WAHA validation
        whatsapp = []
        if validate_whatsapp and self.waha and phone_normalized:
            wa_result = self.waha.validate_for_whatsapp(
                phone_number=phone_normalized,
                country_code=country_code
            )
            if wa_result:
                whatsapp = [wa_result]
                self.stats['waha_validated'] += 1
                logger.debug(f"WAHA validated: {phone_normalized} -> {wa_result}")
        
        # Generate dedupe key and source_link
        dedupe_key = self.generate_dedupe_key(row)
        source_link = self.generate_source_link(dedupe_key)
        
        return {
            'source_link': source_link,
            'partition_key': abs(hash(source_link)) % 32,
            'business_name': str(row.get('name', '')).strip(),
            'business_category': str(row.get('google_business_categories', '')).strip(),
            'business_website': str(row.get('url', '')).strip() if row.get('url') else None,
            'country_code': country_code,
            'city': str(row.get('city', '')).strip(),
            'street': str(row.get('street', '')).strip(),
            'gmaps_phone': phone_normalized,
            'emails': emails,
            'phones': phones,
            'whatsapp': whatsapp,
            'social_facebook': str(row.get('facebook', '')).strip() if row.get('facebook') else None,
            'social_instagram': str(row.get('instagram', '')).strip() if row.get('instagram') else None,
            'scrape_status': 'imported',
        }
    
    def insert_row(self, prepared: Dict[str, Any]) -> bool:
        """Insert new row to database."""
        if self.dry_run:
            return True
        
        conn = self.db.pool.getconn()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO zen_contacts (
                    source_link, partition_key, business_name, business_category, business_website,
                    country_code, city, street, gmaps_phone,
                    emails, emails_count, phones, phones_count, whatsapp, whatsapp_count,
                    social_facebook, social_instagram, scrape_status, created_at, updated_at
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()
                )
                ON CONFLICT (source_link, partition_key) DO NOTHING
            """, (
                prepared['source_link'], prepared['partition_key'],
                prepared['business_name'], prepared['business_category'], prepared['business_website'],
                prepared['country_code'], prepared['city'], prepared['street'], prepared['gmaps_phone'],
                prepared['emails'], len(prepared['emails']),
                prepared['phones'], len(prepared['phones']),
                prepared['whatsapp'], len(prepared['whatsapp']),
                prepared['social_facebook'], prepared['social_instagram'],
                prepared['scrape_status']
            ))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Insert error: {e}")
            conn.rollback()
            return False
        finally:
            self.db.pool.putconn(conn)
    
    def merge_row(self, existing: Dict, prepared: Dict[str, Any]) -> bool:
        """Merge new data into existing row (fill NULLs, merge arrays)."""
        if self.dry_run:
            return True
        
        # Merge arrays
        merged_emails = self.merge_arrays(existing['emails'], prepared['emails'])
        merged_phones = self.merge_arrays(existing['phones'], prepared['phones'])
        merged_whatsapp = self.merge_arrays(existing['whatsapp'], prepared['whatsapp'])
        
        # Fill NULLs
        facebook = existing['social_facebook'] or prepared['social_facebook']
        instagram = existing['social_instagram'] or prepared['social_instagram']
        website = existing['business_website'] or prepared['business_website']
        
        conn = self.db.pool.getconn()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE zen_contacts SET
                    emails = %s, emails_count = %s,
                    phones = %s, phones_count = %s,
                    whatsapp = %s, whatsapp_count = %s,
                    social_facebook = COALESCE(social_facebook, %s),
                    social_instagram = COALESCE(social_instagram, %s),
                    business_website = COALESCE(business_website, %s),
                    updated_at = NOW()
                WHERE source_link = %s
            """, (
                merged_emails, len(merged_emails),
                merged_phones, len(merged_phones),
                merged_whatsapp, len(merged_whatsapp),
                facebook, instagram, website,
                existing['source_link']
            ))
            conn.commit()
            return cursor.rowcount > 0
        except Exception as e:
            logger.error(f"Merge error: {e}")
            conn.rollback()
            return False
        finally:
            self.db.pool.putconn(conn)
    
    def process_row(self, row: Dict[str, Any], validate_whatsapp: bool = True) -> str:
        """Process single row: check existing, insert or merge."""
        self.stats['total_rows'] += 1
        
        name = str(row.get('name', '')).strip()
        street = str(row.get('street', '')).strip()
        city = str(row.get('city', '')).strip()
        country_code = str(row.get('country_code', '')).strip()
        
        if not name:
            self.stats['skipped'] += 1
            return 'skipped'
        
        try:
            # Check existing
            existing = self.check_existing(name, street, city, country_code)
            
            # Prepare data
            prepared = self.prepare_row(row, validate_whatsapp)
            
            if existing:
                # Merge into existing
                if self.merge_row(existing, prepared):
                    self.stats['merged'] += 1
                    return 'merged'
                else:
                    self.stats['errors'] += 1
                    return 'error'
            else:
                # Insert new
                if self.insert_row(prepared):
                    self.stats['inserted'] += 1
                    return 'inserted'
                else:
                    self.stats['errors'] += 1
                    return 'error'
        except Exception as e:
            logger.error(f"Error processing row '{name}': {e}")
            self.stats['errors'] += 1
            return 'error'
    
    def is_data_sheet(self, sheet_name: str, first_row: List) -> bool:
        """Check if sheet contains data (not summary)."""
        sheet_lower = sheet_name.lower()
        
        # Skip summary sheets
        for indicator in self.SUMMARY_INDICATORS:
            if indicator in sheet_lower:
                return False
        
        # Check first row for expected columns
        if first_row:
            first_row_lower = [str(c).lower() for c in first_row]
            return 'name' in first_row_lower or 'phone_number' in first_row_lower
        
        return True
    
    def read_csv(self, file_path: str) -> List[Dict]:
        """Read CSV file."""
        rows = []
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    rows = list(reader)
                logger.info(f"Read {len(rows)} rows from CSV (encoding: {encoding})")
                return rows
            except UnicodeDecodeError:
                continue
        
        raise ValueError(f"Could not read CSV with any encoding: {file_path}")
    
    def read_xlsx(self, file_path: str, sheet_name: Optional[str] = None) -> List[Tuple[str, List[Dict]]]:
        """Read XLSX file, returns list of (sheet_name, rows) tuples."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for XLSX. Run: pip install openpyxl")
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        results = []
        
        sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
        
        for sname in sheets_to_process:
            if sname not in wb.sheetnames:
                logger.warning(f"Sheet '{sname}' not found in workbook")
                continue
            
            ws = wb[sname]
            rows_iter = ws.iter_rows(values_only=True)
            
            # Get header
            try:
                header = next(rows_iter)
                header = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(header)]
            except StopIteration:
                continue
            
            # Check if data sheet
            if not self.is_data_sheet(sname, header):
                logger.info(f"Skipping summary sheet: {sname}")
                continue
            
            # Read rows
            rows = []
            for row_values in rows_iter:
                if not any(row_values):
                    continue
                row_dict = dict(zip(header, row_values))
                rows.append(row_dict)
            
            if rows:
                logger.info(f"Read {len(rows)} rows from sheet '{sname}'")
                results.append((sname, rows))
        
        wb.close()
        return results
    
    def import_file(self, file_path: str, sheet_name: Optional[str] = None, 
                    validate_whatsapp: bool = True) -> Dict[str, int]:
        """Import file to database."""
        file_path = str(file_path)
        ext = Path(file_path).suffix.lower()
        
        if ext == '.csv':
            data_sets = [('csv', self.read_csv(file_path))]
        elif ext in ['.xlsx', '.xls']:
            data_sets = self.read_xlsx(file_path, sheet_name)
        else:
            raise ValueError(f"Unsupported file type: {ext}")
        
        # Process all datasets
        for source_name, rows in data_sets:
            logger.info(f"Processing {source_name}: {len(rows)} rows")
            
            iterator = rows
            if tqdm:
                iterator = tqdm(rows, desc=f"Importing {source_name}", unit="row")
            
            for row in iterator:
                self.process_row(row, validate_whatsapp)
        
        return self.stats


def main():
    parser = argparse.ArgumentParser(
        description='Import CSV/XLSX to PostgreSQL (zen_contacts table)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python toolkit/import_to_db.py result/Argentina.csv
  python toolkit/import_to_db.py result/Report.xlsx --sheet Argentina
  python toolkit/import_to_db.py file.csv --skip-waha --dry-run
        """
    )
    parser.add_argument('file', help='CSV or XLSX file to import')
    parser.add_argument('--sheet', help='Specific sheet name for XLSX (default: all data sheets)')
    parser.add_argument('--skip-waha', action='store_true', help='Skip WAHA WhatsApp validation')
    parser.add_argument('--dry-run', action='store_true', help='Preview without database writes')
    parser.add_argument('--batch-size', type=int, default=100, help='Batch size for commits')
    parser.add_argument('--verbose', '-v', action='store_true', help='Verbose logging')
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    # Validate file
    if not os.path.exists(args.file):
        logger.error(f"File not found: {args.file}")
        sys.exit(1)
    
    # Initialize database
    if args.dry_run:
        logger.info("DRY RUN MODE - No database writes")
        db_writer = None
    else:
        db_writer = create_database_writer(logger)
        if not db_writer:
            logger.error("Failed to create database connection. Check .env file.")
            sys.exit(1)
        
        if not db_writer.connect():
            logger.error("Failed to connect to database")
            sys.exit(1)
    
    # Initialize WAHA validator
    waha_validator = None
    if not args.skip_waha:
        waha_validator = WhatsAppValidator()
        if waha_validator.waha_enabled:
            logger.info(f"WAHA enabled: {waha_validator.waha_base_url}")
        else:
            logger.warning("WAHA not configured - phone validation disabled")
            waha_validator = None
    
    try:
        # Create importer
        importer = DataImporter(
            db_writer=db_writer,
            waha_validator=waha_validator,
            dry_run=args.dry_run,
            batch_size=args.batch_size
        )
        
        # Import file
        stats = importer.import_file(
            file_path=args.file,
            sheet_name=args.sheet,
            validate_whatsapp=not args.skip_waha
        )
        
        # Print results
        print("\n" + "=" * 50)
        print("IMPORT COMPLETE")
        print("=" * 50)
        print(f"Total rows processed: {stats['total_rows']}")
        print(f"  Inserted (new):     {stats['inserted']}")
        print(f"  Merged (existing):  {stats['merged']}")
        print(f"  Skipped:            {stats['skipped']}")
        print(f"  Errors:             {stats['errors']}")
        if not args.skip_waha:
            print(f"  WAHA validated:     {stats['waha_validated']}")
        print("=" * 50)
        
        if args.dry_run:
            print("\n[DRY RUN - No changes made to database]")
        
    finally:
        if db_writer:
            db_writer.close()


if __name__ == '__main__':
    main()
